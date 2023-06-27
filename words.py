import requests
import lxml.html
import openpyxl


def parse(url):
    api = requests.get(url)
    tree = lxml.html.document_fromstring(api.text)
    words = tree.xpath('/html/body/div[4]/div[1]/div[1]/div[1]/div[2]/ul/li/a/text()')
    return words


def main():
    url = "http://www.allscrabblewords.com/{number}-letter-words/"
    i = 2
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    wb.create_sheet(title='Pages')
    while i <= 12:
        words = parse(url.format(number=i))
        sheet = wb['Pages']
        cell = sheet.cell(row=1, column=i - 1)
        cell.value = f'{i} letters'
        for word in words:
            cell = sheet.cell(row=words.index(word)+2, column=i-1)
            cell.value = word
        print(words)
        i += 1
    wb.save('example.xls')


if __name__ == '__main__':
    main()
