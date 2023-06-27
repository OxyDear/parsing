import requests
from bs4 import BeautifulSoup as BS
import json
import lxml.html
import openpyxl


def get_code(url):
    headers = {
        'accept': '*/*',
        'accept-encoding': 'gzip, deflate, br',
        'accept-language': 'ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7',
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36'
    }

    req = requests.get(url=url)
    src = req.text
    print(src)

    with open('funpay.html', 'w', encoding='utf-8') as file:
        file.write(src)


def get_all_folders(url):

    with open('funpay.html', encoding='utf-8') as file:
        src = file.read()

    soup = BS(src, 'lxml')

    # tree = lxml.html.document_fromstring('funpay.html')
    # folders = tree.xpath('/html/body/div[1]/div[1]/section/div[2]/div/div[2]/div/div/div[2]')
    PromoGameListTitle = soup.find(class_='wrapper').find(class_='wrapper-content').find('section').find(class_='content-promo content-promo-index')\
        .find(class_='promo-cd promo-abc').find(class_='promo-games promo-games-all').find(class_='container').find(class_='content-with-cd').find(class_='promo-game-list')\
        .find_all(class_='promo-game-list-title')
    RowRow10Flex = soup.find(class_='wrapper').find(class_='wrapper-content').find('section').find(class_='content-promo content-promo-index')\
        .find(class_='promo-cd promo-abc').find(class_='promo-games promo-games-all').find(class_='container').find(class_='content-with-cd').find(class_='promo-game-list')\
        .find_all(class_='row row-10 flex')

    games = {}
    game_hrefs = []

    for col in RowRow10Flex:
        for r in col.find_all(class_='col-md-3 col-xs-6'):
            for j in r.find(class_='promo-game-item').find('ul').find_all('li'):
                game_hrefs.append(j.find('a').get('href'))
            games[r.find(class_='promo-game-item').find(class_='game-title').find('a').text] = game_hrefs
            game_hrefs = []

    print(games)

    # wb = openpyxl.Workbook()
    # wb.remove(wb['Sheet'])
    # wb.create_sheet(title='Folders')
    #
    # for index, val in enumerate(PromoGameListTitle):
    #     sheet = wb['Folders']
    #     cell = sheet.cell(row=1, column=index + 2)
    #     cell.value = val.text
    # wb.save('commitfirst.xls')


def main():
    get_all_folders('https://funpay.com/')


if __name__ == '__main__':
    main()
